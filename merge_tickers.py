#!/usr/bin/env python3
"""Merge XLSX ticker list with dividend_summary.csv on ticker symbol.

XLSX ticker column:  "1752-HK"  -> stripped to "1752"
CSV symbol column:   "01752"     -> stripped to "1752"
"""
import csv
import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
XLSX_PATH = ROOT / "HKG_Ticker_List_20260618.xlsx"
CSV_PATH  = ROOT / "dividend_summary.csv"
OUT_PATH  = ROOT / "output" / "HKG_Ticker_List_with_Dividends.xlsx"

# --- 1. Load dividend summary into a dict keyed by stripped symbol ---
div_map: dict[str, dict] = {}
with CSV_PATH.open(encoding="utf-8-sig") as f:
    for row in csv.DictReader(f):
        key = row["symbol"].lstrip("0")  # "01752" -> "1752"
        div_map[key] = row

# --- 2. Load XLSX ---
wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb.active

# --- 3. Insert new columns after column J (last existing column) ---
# New headers
new_headers = ["has_dividend_csv", "dividend_row_count", "dividend_csv_path"]
insert_col  = ws.max_column + 1   # column K

for col_idx, header in enumerate(new_headers, start=insert_col):
    ws.cell(row=1, column=col_idx).value = header

# --- 4. Fill in dividend data per row ---
matched = not_matched = 0
for row in ws.iter_rows(min_row=2, values_only=False):
    ticker_cell = row[4]  # column E = ticker ("1752-HK")
    ticker_val  = ticker_cell.value
    if not ticker_val:
        for col_idx, header in zip(
            range(insert_col, insert_col + 3), new_headers
        ):
            row[0].parent.cell(row=row[0].row, column=col_idx).value = ""
        continue

    # Strip "-HK" suffix
    key = str(ticker_val).replace("-HK", "").lstrip("0")
    if key == "":
        key = "0"

    div_row = div_map.get(key, {})

    # CSV keys: "has_dividend_csv", "row_count", "csv_path"
    # Map to output column names
    csv_to_output = {
        "has_dividend_csv":  "has_dividend_csv",
        "row_count":         "dividend_row_count",
        "csv_path":          "dividend_csv_path",
    }

    for col_offset, csv_key in enumerate(csv_to_output):
        cell = row[0].parent.cell(row=row[0].row, column=insert_col + col_offset)
        cell.value = div_row.get(csv_key, "")

    if div_row:
        matched += 1
    else:
        not_matched += 1

# --- 5. Save ---
OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT_PATH)

print(f"Saved : {OUT_PATH}")
print(f"Total XLSX rows : {ws.max_row - 1}")
print(f"Matched dividend: {matched}")
print(f"No dividend data: {not_matched}")
